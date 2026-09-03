#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
处理 STR 优化三批数据的「样本 ↔ barcode ↔ RUN」xlsx，并与盘上数据目录交叉验证。

输入（3 个 xlsx，两种表结构）：
  第一批-样本对应barcode.xlsx       : 样本名称, barcode                     （无 RUN 号）
  第二批-样本对应barcode.xlsx       : 质粒名称, 对应barcode, DPN名称, RUN号, 服务器
  第三批-样本对应barcode 30-43.xlsx : 同上
  第二/三批中 RUN号/DPN/服务器 只在每个 block 的首行给出，其余行留空 → 按块向下填充。

barcode 三层命名（映射表来自 数据介绍.md）：
  24标签-N   (xlsx 里的标签位置)
  barcode-NN (新编号，测序人员可感知)
  Adaptor-barcodeXXX-Y (原编号，部分 run 的 fastq 文件名用这个)

输出（写入脚本所在目录）：
  样本-run-barcode对应关系.tsv / .xlsx   统一平铺表
  xlsx处理报告.md                        每批 block 汇总 + 数据文件核对 + 未分配/池外 barcode
"""
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path("/data1/ccs_data/str-optimization")
OUTDIR = Path(__file__).resolve().parent

# 24 标签位置 -> (barcode 新编号, 原编号)，来自 数据介绍.md 的映射表
POS_TO_BARCODE = {
    1: ("barcode-01", "Adaptor-barcode251-4"),
    2: ("barcode-02", "Adaptor-barcode250-4"),
    3: ("barcode-03", "Adaptor-barcode271-3"),
    4: ("barcode-04", "Adaptor-barcode266-2"),
    5: ("barcode-05", "Adaptor-barcode229-0"),
    6: ("barcode-06", "Adaptor-barcode255-3"),
    7: ("barcode-07", "Adaptor-barcode261-3"),
    8: ("barcode-08", "Adaptor-barcode206-3"),
    9: ("barcode-09", "Adaptor-barcode230-2"),
    10: ("barcode-10", "Adaptor-barcode256-3"),
    11: ("barcode-11", "Adaptor-barcode207-4"),
    12: ("barcode-12", "Adaptor-barcode222-2"),
    13: ("barcode-13", "Adaptor-barcode270-0"),
    14: ("barcode-14", "Adaptor-barcode201-5"),
    15: ("barcode-15", "Adaptor-barcode253-4"),
    16: ("barcode-16", "Adaptor-barcode217-2"),
    17: ("barcode-17", "Adaptor-barcode214-1"),
    18: ("barcode-18", "Adaptor-barcode221-3"),
    19: ("barcode-19", "Adaptor-barcode216-0"),
    20: ("barcode-20", "Adaptor-barcode254-5"),
    21: ("barcode-21", "Adaptor-barcode220-0"),
    22: ("barcode-22", "Adaptor-barcode225-1"),
    23: ("barcode-23", "Adaptor-barcode233-0"),
    24: ("barcode-24", "Adaptor-barcode239-2"),
}
ORIGINAL_TO_POS = {orig: pos for pos, (_, orig) in POS_TO_BARCODE.items()}

BATCHES = [
    # 第一批 xlsx 无 RUN 号；经确认只用最新 run（其余 run 目录不再展开）
    {"batch": "第一批", "dir": "first-batch-of-data", "xlsx": "第一批-样本对应barcode.xlsx",
     "runs_keep": ["20260723_250302Y0001_Run0003"],
     "merged": "STR第一批一代测序/STR一代测序/merged_output"},
    {"batch": "第二批", "dir": "second-batch-of-data", "xlsx": "第二批-样本对应barcode.xlsx",
     "merged": "sanger_result/merged_output"},
    {"batch": "第三批", "dir": "third-batch-of-data", "xlsx": "第三批-样本对应barcode 30-43.xlsx",
     "merged": "STR第三批一代测序/STR第三批一代测序/merged_output"},
]

RUN_DIR_RE = re.compile(r"^\d{8}_.*_Run\d+$")
BARCODE_SUBDIRS = ("barcode_assign", "barcodes_reads_fastq_amplicon")


def norm(v):
    return str(v).strip() if v is not None else ""


def parse_pos(barcode_str):
    """'样本-24标签-1' / '24标签-1' -> 1"""
    m = re.search(r"标签-(\d+)", barcode_str or "")
    return int(m.group(1)) if m else None


def read_xlsx(path):
    """解析 xlsx，返回 (records, warnings)。block 级字段(RUN号/DPN/服务器)向下填充。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [norm(c) for c in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_sample = col("样本名称", "质粒名称")
    i_barcode = col("barcode", "对应barcode")
    i_run = col("RUN号", "Run号", "run号")
    i_dpn = col("DPN名称")
    i_srv = col("服务器")
    if i_sample is None or i_barcode is None:
        raise ValueError(f"{path}: 未找到 样本/barcode 列, header={header}")

    warnings = []
    records = []
    block = {"run": "", "dpn": "", "server": ""}
    for ln, r in enumerate(rows[1:], start=2):
        if r is None or all(v is None or norm(v) == "" for v in r):
            continue
        run = norm(r[i_run]) if i_run is not None else ""
        if run:  # RUN 号非空 → 新 block，DPN/服务器 随块更新
            block = {
                "run": run,
                "dpn": norm(r[i_dpn]) if i_dpn is not None else "",
                "server": norm(r[i_srv]) if i_srv is not None else "",
            }
        sample = norm(r[i_sample])
        barcode = norm(r[i_barcode])
        if not sample and not barcode:
            continue
        pos = parse_pos(barcode)
        if pos is None:
            warnings.append(f"第{ln}行 barcode 无法解析位置: {barcode!r}")
            continue
        records.append({
            "sample": sample, "barcode_raw": barcode, "pos": pos,
            "run": block["run"], "dpn": block["dpn"], "server": block["server"],
        })
    return records, warnings


def find_barcode_file(run_dir, pos):
    """在不同命名方案下找该位置对应的 fastq。"""
    _, orig = POS_TO_BARCODE[pos]
    candidates = [
        run_dir / "barcode_assign" / f"Barcode{pos:02d}.fastq",          # 第一/二批部分 run
        run_dir / "barcode_assign" / f"{orig}.fastq",                    # 第三批
        run_dir / "barcodes_reads_fastq_amplicon" / f"{orig}.fastq",     # 第二批 Run0001/0002
    ]
    for c in candidates:
        if c.is_file():
            return c
    return None


def count_reads(fastq_path):
    with open(fastq_path, "r") as f:
        n = sum(1 for _ in f)
    return n // 4


def scan_run_barcodes(run_dir):
    """列出 run 目录下所有 per-barcode fastq。
    返回 (pool: {pos: (name, reads)}, offpool: [(name, reads)])
    """
    pool, offpool = {}, []
    for d in BARCODE_SUBDIRS:
        dd = run_dir / d
        if not dd.is_dir():
            continue
        for f in sorted(dd.glob("*.fastq")):
            stem = f.stem
            m = re.match(r"^Barcode(\d+)$", stem)
            pos = int(m.group(1)) if m else ORIGINAL_TO_POS.get(stem)
            if pos is not None and pos not in pool:
                pool[pos] = (stem, count_reads(f))
            else:
                offpool.append((stem, count_reads(f)))
    return pool, offpool


def seq_len(path, header_pat=None):
    """文件序列总长（bp）；header_pat 给定时只统计 header 匹配的记录。"""
    total, cur_match = 0, False
    with open(path, encoding="utf-8") as f:
        for line in f:
            if line.startswith(">"):
                cur_match = (header_pat is None) or bool(header_pat.match(line[1:].strip()))
            elif cur_match:
                total += len(line.strip())
    return total


def sample_token(sample):
    """样本名去掉 STR 前缀：'STR1-1'->'1-1'，'3N-1'->'3N-1'"""
    return re.sub(r"^STR", "", sample)


def find_reference(sample, merged_dir):
    """找样本的参考序列：优先独立文件，其次汇总文件（intersect.fa / merged.fa）里的记录。
    返回 (path, note, header_pat)。"""
    tok = sample_token(sample)
    pat = re.compile(r"^.*STR" + re.escape(tok) + r"\.(?:intersect\.)?fa$")
    hdr_pat = re.compile(r"^.*STR" + re.escape(tok) + r"(\s|$)")
    if merged_dir is None or not merged_dir.is_dir():
        return None, "参考序列目录不存在", None
    hits = sorted(f for f in merged_dir.iterdir() if f.is_file() and pat.match(f.name))
    if hits:
        return hits[0], "", None
    for agg in ("intersect.fa", "merged.fa"):
        p = merged_dir / agg
        if p.is_file():
            with open(p, encoding="utf-8") as f:
                if any(line.startswith(">") and hdr_pat.match(line[1:].strip()) for line in f):
                    return p, "无独立参考文件，取汇总文件内记录", hdr_pat
    return None, "未找到参考序列", None


def main():
    all_rows = []
    report = []
    header = ["批次", "样本名", "24标签", "barcode新编号", "barcode原编号",
              "RUN号", "DPN名称", "服务器", "fastq文件", "reads数",
              "参考序列", "参考长度bp", "备注"]

    for b in BATCHES:
        bdir = BASE / b["dir"]
        xlsx_path = bdir / b["xlsx"]
        records, warnings = read_xlsx(xlsx_path)
        run_dirs = sorted(d for d in bdir.iterdir() if d.is_dir() and RUN_DIR_RE.match(d.name))
        run_dirs_by_name = {d.name: d for d in run_dirs}

        report.append(f"## {b['batch']}  (`{b['dir']}/`)")
        report.append(f"- xlsx：`{xlsx_path.name}`，共 {len(records)} 行样本记录")
        report.append(f"- 盘上 run 目录：{', '.join(d.name for d in run_dirs) or '（无）'}")
        report.append("- block 划分（RUN 号所在行起为新 block）：")

        blocks = []
        for r in records:
            key = (r["run"], r["dpn"], r["server"])
            if not blocks or blocks[-1]["key"] != key:
                blocks.append({"key": key, "n": 0})
            blocks[-1]["n"] += 1
        for blk in blocks:
            run, dpn, srv = blk["key"]
            run_label = run or "（xlsx 未标注 RUN 号）"
            report.append(f"  - {run_label} | DPN={dpn or '-'} | 服务器={srv or '-'} | {blk['n']} 个样本")
        if warnings:
            report.append(f"- xlsx 解析警告：{warnings}")

        # 参考序列（一代测序合并结果）：样本名 -> 文件
        merged_dir = BASE / b["dir"] / b["merged"] if b.get("merged") else None
        ref_cache = {}

        def get_ref(sample):
            if sample not in ref_cache:
                path, note, hdr = find_reference(sample, merged_dir)
                ref_cache[sample] = (path, note, seq_len(path, hdr) if path else 0)
            return ref_cache[sample]

        ref_stat = {"独立文件": 0, "汇总文件": 0, "缺失": 0}
        for s in sorted({r["sample"] for r in records}):
            path, note, _ = get_ref(s)
            ref_stat["缺失" if not path else ("汇总文件" if note else "独立文件")] += 1
        report.append(
            f"- 参考序列目录：`{b['dir']}/{b['merged']}`"
            f"（独立文件 {ref_stat['独立文件']} / 汇总文件内 {ref_stat['汇总文件']} / 缺失 {ref_stat['缺失']}）"
        )
        if ref_stat["缺失"]:
            miss = sorted({r["sample"] for r in records if not get_ref(r["sample"])[0]})
            report.append(f"- **缺失参考序列**：{', '.join(miss)}")
        # 反向核对：有参考序列但 xlsx 无对应样本
        sample_toks = {sample_token(r["sample"]) for r in records}
        orphans = []
        if merged_dir and merged_dir.is_dir():
            for f in sorted(merged_dir.iterdir()):
                if not f.is_file():
                    continue
                m = re.match(r"^.*STR(.+?)\.(?:intersect\.)?fa$", f.name)
                if m:
                    orphans.append((f.name, m.group(1)))
            for agg in ("intersect.fa", "merged.fa"):
                p = merged_dir / agg
                if not p.is_file():
                    continue
                for line in open(p, encoding="utf-8"):
                    m = re.match(r"^>(\S+)", line)
                    if m:
                        mm = re.search(r"STR(.+)$", m.group(1))
                        if mm and mm.group(1) not in {t for _, t in orphans}:
                            orphans.append((f"{agg}::{m.group(1)}", mm.group(1)))
        orphans = [o for o in orphans if o[1] not in sample_toks]
        if orphans:
            report.append("- 有参考序列但 ccs xlsx 无对应样本：" +
                          ", ".join(f"{n}（{t}）" for n, t in sorted(orphans)))

        # 展开成 (样本, run, pos) 行；第一批 xlsx 无 RUN 号 → 按 runs_keep 指定的 run 展开
        keep = b.get("runs_keep")
        if keep:
            report.append(f"- xlsx 无 RUN 号，经确认仅采用：{', '.join(keep)}")
        expanded = {}
        for r in records:
            if r["run"]:
                if r["run"] not in run_dirs_by_name:
                    report.append(f"- **异常**：xlsx 中的 RUN 号 `{r['run']}` 在盘上找不到目录")
                expanded[(r["sample"], r["run"], r["pos"])] = {**r, "expanded": False}
            else:
                for d in run_dirs:
                    if keep and d.name not in keep:
                        continue
                    expanded[(r["sample"], d.name, r["pos"])] = {**r, "run": d.name, "expanded": True}

        # 逐 run 核对数据文件 + 找未分配/池外 barcode
        used_by_run = {}
        for (sample, run, pos) in expanded:
            used_by_run.setdefault(run, set()).add(pos)

        for run in sorted(used_by_run):
            run_dir = run_dirs_by_name.get(run)
            used = used_by_run[run]
            pool, offpool = scan_run_barcodes(run_dir) if run_dir else ({}, [])

            for pos in sorted(used):
                sample = next(s for (s, rn, p) in expanded if rn == run and p == pos)
                rec = expanded[(sample, run, pos)]
                new_bc, orig = POS_TO_BARCODE[pos]
                f = find_barcode_file(run_dir, pos) if run_dir else None
                reads = count_reads(f) if f else 0
                ref_path, ref_note, _ = get_ref(sample)
                ref_len = get_ref(sample)[2]
                note = "xlsx 未标注 RUN 号，按确认的最新 run 展开" if rec["expanded"] else ""
                if not f:
                    note = "；".join(x for x in [note, "数据文件缺失"] if x)
                if ref_note:
                    note = "；".join(x for x in [note, ref_note] if x)
                all_rows.append([b["batch"], sample, f"24标签-{pos}", new_bc, orig,
                                 run, rec["dpn"], rec["server"],
                                 str(f) if f else "", reads,
                                 str(ref_path) if ref_path else "", ref_len, note])

            report.append(f"### run: `{run}`")
            missing = [f"24标签-{p}" for p in sorted(used)
                       if not (run_dir and find_barcode_file(run_dir, p))]
            if missing:
                report.append(f"- **缺失**：{', '.join(missing)}（xlsx 有、盘上找不到 fastq）")
            unassigned = [p for p in sorted(pool) if p not in used]
            if unassigned:
                items = ", ".join(f"24标签-{p}({pool[p][0]}, {pool[p][1]}reads)" for p in unassigned)
                report.append(f"- 盘上未分配（池内位置未用于本 run 的 block）：{items}")
            offpool_with_reads = [(n, r) for n, r in offpool if r > 0]
            offpool_empty = len(offpool) - len(offpool_with_reads)
            if offpool_with_reads:
                items = ", ".join(f"`{n}`({r}reads)" for n, r in sorted(offpool_with_reads))
                report.append(f"- 池外 barcode 有 reads（疑似邻码误分）：{items}")
            if offpool_empty:
                report.append(f"- 池外空 fastq（0 reads，忽略）：{offpool_empty} 个")
            report.append("")

    # ---- 写 TSV / xlsx ----
    batch_order = {b["batch"]: i for i, b in enumerate(BATCHES)}
    all_rows.sort(key=lambda x: (batch_order[x[0]], x[5], int(re.search(r"标签-(\d+)", x[2]).group(1))))
    tsv_path = OUTDIR / "样本-run-barcode对应关系.tsv"
    with open(tsv_path, "w", encoding="utf-8") as f:
        f.write("\t".join(header) + "\n")
        for row in all_rows:
            f.write("\t".join(str(c) for c in row) + "\n")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "样本-run-barcode"
    ws.append(header)
    for row in all_rows:
        ws.append(row)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for i, w in enumerate([8, 10, 10, 14, 24, 32, 10, 12, 72, 9, 72, 12, 36], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    xlsx_out = OUTDIR / "样本-run-barcode对应关系.xlsx"
    wb.save(xlsx_out)

    # ---- 报告 ----
    report = [
        "# 三批 xlsx（样本 ↔ barcode ↔ RUN）处理报告",
        "",
        f"- 统一表：`{tsv_path.name}` / `{xlsx_out.name}`（共 {len(all_rows)} 行）",
        "- barcode 映射（24标签 ↔ 新编号 ↔ 原编号）来自 数据介绍.md 的映射表",
        "",
    ] + report
    report_path = OUTDIR / "xlsx处理报告.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(report) + "\n")

    print(f"OK: {len(all_rows)} rows -> {tsv_path.name}, {xlsx_out.name}")
    print(f"report -> {report_path.name}")


if __name__ == "__main__":
    sys.exit(main())
